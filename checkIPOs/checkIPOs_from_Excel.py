# -*- coding: utf-8 -*-
"""
Created on Thu Apr  6 21:16:14 2023

@author: alexander.busch@alumni.ntnu.no

Files: https://incose2.sharepoint.com/sites/SEHv5GermanTranslation >> SEHv5 >> Übersetzungsfiles >> Abbildungen >> IPO_consistency_check

"""


# -----------------------------------------------------------------------------
# LIBRARIES
# -----------------------------------------------------------------------------

import os
import openpyxl


"""
-----------------------------------------------------------------------------
READ DATA FROM .xlsx
-----------------------------------------------------------------------------
"""

# Directory path
path = r"C:\Users\alexanderb\INCOSE\SEHv5 German Translation - SEHv5\_Übersetzungsfiles\Abbildungen"
path = r'/home/trbprnz/Downloads/'
path = r'/home/trbprnz/cloud/OneDrive/INCOSE_SEHv5/SEHv5/_Übersetzungsfiles/Abbildungen'

# pptx file containing the IPO diagrams
file = r'2023-03-10 SEH5E-IPOs-Descr-N2_copy.xlsx'
file = r'2023-03-10 SEH5E-IPOs-Descr-N2.xlsx'

# Load the Excel workbook
workbook = openpyxl.load_workbook(os.path.join(path,file))

# Select the worksheet to read from
worksheet = workbook['A - IPO Diagram Content']

# Define the range of cells to read from
#text=['A1:D', str(worksheet.max_row)]
#range = ''.join(text)

# Use a list comprehension to read the values in the cell range
#data = [cell[0].value for cell in worksheet[range]]

# Print the cell values
#print(pptxpresentations)
#print(pptxfiles)

# Create an empty list to hold the data
IPO_data = []

# Loop through the rows and columns of the selected range
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=6, values_only=True):
    # Create a list to hold the values for this row
    row_data = []
    for cell in row:
        # If the cell contains multiple lines, split the text into a list
        if isinstance(cell, str) and '\n' in cell:
            cell = cell.split('\n')
            cell = sorted(cell)
        row_data.append(cell)
    IPO_data.append(row_data)



""" 
-----------------------------------------------------------------------------
CLEAN-UP DATA
-----------------------------------------------------------------------------
"""

# Combine string in column title that was erroneously imported as two separate list entries
IPO_data[0][4]=' '.join(IPO_data[0][4])

# Remove the first ('ID') and third ('INCOSE SE Handbook 4E 2015 Inputs') columns from the table
for row in IPO_data:
    row.pop(0)
    row.pop(1)

# Extract original header and remove first row in data
header_original = IPO_data[0]
IPO_data.pop(0)

# Relabel row 'External (...)' with 'External'
IPO_data[0][0]=IPO_data[0][0][0]

# Extract the data for the following process model variables
external_input= sorted(IPO_data[0][1])
external_output= sorted(IPO_data[0][3])
controls = sorted(IPO_data[1][1])
enablers = sorted(IPO_data[2][1])
processes = [row[0] for row in IPO_data]


# Remove rows at indices 0, 1, 2 from the data
indices_to_remove = [0, 1, 2]
for index in sorted(indices_to_remove, reverse=True):
    IPO_data.pop(index)
    
# Define the header for the table
headers = ['Process',
           'Inputs',
           'Activities',
           'Outputs']




""" 
-----------------------------------------------------------------------------
Create Excel file to visualize data
-----------------------------------------------------------------------------
"""

def write_to_excel(IPO_data, filename, headers):
    
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    
    if os.path.exists(filename):
        os.remove(filename)
        
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # Headers
    sheet.append(headers)    
    
    # Data
    row_start = 2
    
    for row_idx, row in enumerate(IPO_data):
        for col_idx, col in enumerate(row):
            if col is None:
                sheet.cell(row=row_idx+row_start, column=col_idx+1).value = ""
            elif isinstance(col, str):
                sheet.cell(row=row_idx+row_start, column=col_idx+1).value = col
            else:
                #for i, subcol in enumerate(col):
                 #   sheet.cell(row=row_idx+1+i, column=col_idx+1).value = subcol
                 cell_value = '\n'.join(col)
                 sheet.cell(row=row_idx+row_start, column=col_idx+1).value = cell_value
                 
     # Get the number of rows and columns in the worksheet
    num_rows = sheet.max_row
    num_cols = sheet.max_column
                 
    # Create a new Table object and set its style
    table = Table(displayName="IPO", ref=':'.join(['A1',get_column_letter(num_cols) + str(num_rows)]))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    # Autofit all columns
    for col in sheet.columns:
        max_length = 0
        for cell in col:
            try:
                cell_value = str(cell.value)
                lines = cell_value.split('\n')
                for line in lines:
                    if len(line) > max_length:
                        max_length = len(line)
            except:
                pass
        adjusted_width = max_length
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width
        
    # Autofit all rows
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                cell_value = str(cell.value)
                cell_text = openpyxl.utils.cell.text.get_column_letter(cell.column) + str(cell.row)
                lines = cell_value.split('\n')
                font_size = cell.font.size
                for line in lines:
                    height = openpyxl.utils.units.points_to_pixels(openpyxl.utils.units.inches_to_points(font_size*0.75))
                    wrapped_text = openpyxl.utils.text._wrap_text(line, font_size, adjusted_width)
                    line_height = height * len(wrapped_text)
                    if line_height > max_height:
                        max_height = line_height
            except:
                pass
        if max_height > 0:
            sheet.row_dimensions[row[0].row].height = max_height
    
    # Activate wrap text
    for row in sheet.rows:
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        
    # Add the Table object to the worksheet and save the workbook to a file
    wb.save(filename)


# Create Excel file for entire IPO_data set
filename=os.path.join(path,"IPO_consistency_check","IPO_0.xlsx")
write_to_excel(IPO_data, filename, headers)

# Create Excel file for benchmark with Rev 2023-04
filename=os.path.join(path,"IPO_consistency_check","Rev2023-03.xlsx")
data_to_excel = [sorted(processes), sorted(external_input), sorted(external_output)]

# Find the length of the longest sublist
max_len = max(len(sublist) for sublist in data_to_excel)

# Create a list of tuples, where each tuple contains the i-th element
# of each sublist (padded with None if necessary)
transposed_list = []
for i in range(max_len):
    tuple_ = tuple(sublist[i] if i < len(sublist) else None for sublist in data_to_excel)
    transposed_list.append(tuple_)

write_to_excel(transposed_list, filename, ['Processes', 'Ext. inputs', 'Ext. outputs'])





""" 
-----------------------------------------------------------------------------
# Create unique lists for Inputs & Outputs
-----------------------------------------------------------------------------
"""

# Extract unique strings for inputs & outputs
def extract_unique_arrays(data,column):
    
    # Convert table data to array
    if column>0:
        # Initialize an empty list to store all arrays
        returnarray = []
                
        # Loop through each row in the table
        for row in data:
            # Get the array from the specified column and append it to the returnarray list
            array = row[column]
            if array is not None:
                if isinstance(array, str):
                    returnarray.append([array])
                else:
                    returnarray.append(array)
        
        # Flatten the list of arrays into a single list
        returnarray = [item for sublist in returnarray for item in sublist]
    
    else:
        returnarray = data
    
    # Remove empty cells
    returnarray = [x for x in returnarray if x != '']
    
    # Remove duplicate entries
    returnarray = list(set(returnarray))
    
    # Sort
    returnarray = sorted(returnarray)
    
    return returnarray

inputs_unique = extract_unique_arrays(IPO_data,1)
outputs_unique = extract_unique_arrays(IPO_data,3)

#print(inputs_unique)


""" 
-----------------------------------------------------------------------------
# Check the similarity of the individual strings in the categorical lists
-----------------------------------------------------------------------------
"""

def find_similar_strings(strings_list,treshold):
    """
    This function takes a list of strings and compares each string with all the other strings in the list regarding similarities. 
    It uses SequenceMatcher from difflib module to compare the similarity between two strings.
    https://docs.python.org/2/library/difflib.html
    The similarity metric is a percentage defined by the variable treshold.
    The function returns a list of lists containing the similar strings.
    """
    
    from difflib import SequenceMatcher, get_close_matches
    
    # Initialize an empty list to store the similar strings
    similar_strings = []
    #good_enough_matches = []    

    # Loop through each string in the list
    for i, string1 in enumerate(strings_list):
        
        # Return a list of the best “good enough” matches
        # Not used as it produces way too many results
        #good_enough_matches.append(get_close_matches(string1, strings_list, 3, treshold))



        # Initialize an empty list to store the strings that are similar to the current string
        similar_strings_i = []

        # Loop through all the strings in the list again
        for j, string2 in enumerate(strings_list):

            # Skip comparing the string with itself
            if i == j:
                continue

            # Use SequenceMatcher to compare the similarity between the two strings
            similarity_ratio = SequenceMatcher(None, string1, string2).ratio()            

            # If the similarity ratio is above a certain threshold, add the string to the similar_strings_i list
            if similarity_ratio > treshold:
                similar_strings_i.append(string2)

        # Add the similar_strings_i list to the similar_strings list
        similar_strings.append(similar_strings_i)

    # Return the final similar_strings list
    #return good_enough_matches
    return similar_strings

inputs_similar_strings = find_similar_strings(inputs_unique,treshold=0.95)
outputs_similar_strings = find_similar_strings(outputs_unique,treshold=0.95)

# Print the final similar_strings list
print('Similar strings in Inputs')
print([sublist for sublist in inputs_similar_strings if sublist])
print()
print('Similar strings in Outputs')
print([sublist for sublist in outputs_similar_strings if sublist])
print()


# Define a dictionary of replacements based on previous results
replacements = {
    "Maintenance and logistics report": "Maintenance and logistic report",
    "xxx": "yyy",
    "zzz": "www"}
# logistics ist richtig


# Loop through each element in the IPO_data list,
# check the values against the dictionary and update in case needed
for i in range(len(IPO_data)):
    for j in range(len(IPO_data[i])):
        # Check if the current element is a string that needs to be replaced
        # Check if the current element is a list
        if isinstance(IPO_data[i][j], list):
            # Loop through each string in the list and replace as needed
            for k in range(len(IPO_data[i][j])):
                if isinstance(IPO_data[i][j][k], str) and IPO_data[i][j][k] in replacements:
                    IPO_data[i][j][k] = replacements[IPO_data[i][j][k]]
                    print('Replaced with')
                    print(IPO_data[i][j][k])
                    print()
        # Otherwise, check if the current element is a string that needs to be replaced
        elif isinstance(IPO_data[i][j], str) and IPO_data[i][j] in replacements:
            IPO_data[i][j] = replacements[IPO_data[i][j]]
            print('Replaced with')
            print(IPO_data[i][j])
            print()


# Update the categorical arrays
inputs_unique = extract_unique_arrays(IPO_data,1)
outputs_unique = extract_unique_arrays(IPO_data,3)

# Create Excel file
filename=os.path.join(path,"IPO_consistency_check","IPO_1.xlsx")
write_to_excel(IPO_data, filename, headers)


"""
-----------------------------------------------------------------------------

Remove all outputs from outputs_unique for which a corresponding input exists in inputs unique and vice versa

-----------------------------------------------------------------------------
"""

# Convert the lists to sets
set1 = set(inputs_unique)
set2 = set(outputs_unique)

# Clean the lists from items occuring in either list
inputs_unique_cleaned = list(set1.difference(set2))
outputs_unique_cleaned = list(set2.difference(set1))

sorted(outputs_unique_cleaned)


# Find the unique strings between set1 and set2
# unique_strings = set1.difference(set2).union(set2.difference(set1))
# print(unique_strings)



""" 
-----------------------------------------------------------------------------
# Bundle individual outputs into categories, such that these can be validated against the inputs

In inputs_unique
'Records/artifacts'
'Project procedures'
'Project reports'
'Project strategies/approaches'

In outputs_unique
'... records/artifacts'
'... procedure', 'Organization procedures'
'... report'
'... strategy/approach'


Organization procedures
Project procedures

-----------------------------------------------------------------------------
"""

# Define a dictionary of replacements based on David's input and manual analysis
replacements = {
    "records/artifacts": "Records/artifacts",
    "procedure": "Project procedures",
    "report": "Project reports",
    "strategy/approach": "Project strategies/approaches"}

def replace_with_generic_category(strings_list, replacements):
    new_list = []
    
    for string in strings_list:
        for key, value in replacements.items():
            if key in string:
                string = value
        new_list.append(string)
    
    return new_list

outputs_unique_categorized = extract_unique_arrays(replace_with_generic_category(outputs_unique_cleaned, replacements),0)





CHECK


replacements = {
    "records/artifacts": "Records/artifacts",
    "procedure": "Project procedures",
    "report": "Project reports",
    "strategy/approach": "Project strategies/approaches"
}

# Add exceptions
exceptions = {
    "Records/artifacts/reports": "records/artifacts/reports",
    "Project procedures/reporting": "procedure/reporting"
}

def replace_with_generic_category(strings_list, replacements, exceptions):
    new_list = []

    for string in strings_list:
        # Check for exceptions first
        if string in exceptions:
            new_list.append(exceptions[string])
        else:
            for key, value in replacements.items():
                if key in string:
                    string = value
            new_list.append(string)

    return new_list








""" 
-----------------------------------------------------------------------------
# Check everything (inputs, outputs, ...) against termbase
-----------------------------------------------------------------------------
"""









""" 
-----------------------------------------------------------------------------
# Validate all inputs that are not outputs against external outputs
-----------------------------------------------------------------------------
"""

# Find strings in inputs that are not in outputs
diff = set(inputs_unique_cleaned) - set(outputs_unique_categorized)
print("Strings in process inputs but not in process outputs:", sorted(diff))
# By definition, these should be all part of external_ouput
missing = set(diff)-set(external_output)
print("Strings missing in external_inputs:", sorted(missing))


""" 
-----------------------------------------------------------------------------
# Validate all outputs that are not inputs against external intputs
-----------------------------------------------------------------------------
"""

# Find strings in process outputs that are not in process inputs
diff2 = set(outputs_unique_categorized) - set(inputs_unique_cleaned)
print("Strings in outputs but not in inputs:", sorted(diff2))
# By definition, these should be all part of external_input
missing = set(diff2) - set(external_input)
print("Strings missing in external_inputs:", sorted(missing))



""" 
Findings

1. Not all categories defined above are fully correct, e.g.
"Supply strategy/approach" is not part of "Project strategies/approaches", instead it is output of SUP and input of PM
"Organization tailoring strategy/approach" is output of TLR and input to LCMM

2. Some categorizations are missing, e.g. 
"Measurement data" is input to "MEAS" but appears as output
- "Critical performance measurement data" in BMA, SNRD, SRD, SAD, DD
- "Organizational measurement data" in LCMM
- "Project measurement data" in PAC

"""
